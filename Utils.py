from pathlib import Path
import customtkinter
import openpyxl.drawing.text
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, Paragraph
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
import shutil
import pandas as pd
from openpyxl.chart.text import RichText
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


class ScrollableCheckBoxFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, item_list, command=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        self.checkbox_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        checkbox = customtkinter.CTkCheckBox(self, text=reverse_hebrew_sentence(item))
        if self.command is not None:
            checkbox.configure(command=self.command)
        checkbox.grid(row=len(self.checkbox_list), column=0, pady=(0, 10))
        self.checkbox_list.append(checkbox)

    def remove_item(self, item):
        for checkbox in self.checkbox_list:
            if item == checkbox.cget("text"):
                checkbox.destroy()
                self.checkbox_list.remove(checkbox)
                return

    def get_checked_items(self):
        return [reverse_hebrew_sentence(checkbox.cget("text")) for checkbox in self.checkbox_list if
                checkbox.get() == 1]


class ScrollableRadiobuttonFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, item_list, command=None, text_variable=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        if text_variable is not None:
            self.radiobutton_variable = text_variable
        else:
            self.radiobutton_variable = customtkinter.StringVar()
            if len(item_list) > 0:
                self.radiobutton_variable.set(item_list[0])

        self.radiobutton_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        radiobutton = customtkinter.CTkRadioButton(self, text=reverse_hebrew_sentence(item), value=item,
                                                   variable=self.radiobutton_variable)
        if self.command is not None:
            radiobutton.configure(command=self.command)
        radiobutton.grid(row=len(self.radiobutton_list), column=0, pady=(0, 10))
        self.radiobutton_list.append(radiobutton)

    def remove_item(self, item):
        for radiobutton in self.radiobutton_list:
            if item == radiobutton.cget("text"):
                radiobutton.destroy()
                self.radiobutton_list.remove(radiobutton)
                return

    def get_checked_item(self):
        return self.radiobutton_variable.get()


def reverse_hebrew_sentence(sentence):
    split = sentence.split(' ')
    return " ".join(split[::-1])


def custom_sort_key(val):
    if val == 'UTC':
        return 0, ''  # Make "UTC" come first
    try:
        return 1, int(val)  # Then sort by numeric values
    except ValueError:
        return 2, str(val)  # Fallback to string comparison for other values


def generate_new_file(filename):
    path = Path(filename)
    output_file = path.with_stem(f"{path.stem}(app_generated)")
    if not Path(output_file).is_file():
        output_path = shutil.copy(filename, output_file)
    else:
        output_path = output_file
    return output_path


def append_df_to_excel(filename, df, sheet_name='Sheet1',start_distance=1):
    # Try to open an existing workbook
    try:
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.workbook = book
        if sheet_name in writer.workbook.sheetnames:
            startrow = writer.workbook[sheet_name].max_row
        else:
            startrow = 0
    except FileNotFoundError:
        # File does not exist yet, create a new one
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w')
        startrow = 0
    # Write the dataframe to the Excel file
    if startrow == 0:
        start_distance = 2
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow + start_distance, index=False, header=True)
    writer.close()
    # Reload the workbook and the sheet to apply formatting
    book = load_workbook(filename)
    sheet = book[sheet_name]

    # Apply font and borders
    font = Font(name='David', size=12)
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))
    alignment = Alignment(horizontal='center')
    dot_index = 0
    if "DOT" in df.columns:
        dot_index = list(df.columns).index("DOT") + 1
    frequency = False
    for col in df.columns:
        if type(col) is str and "שכיחות" in col:
            frequency = True
    # Find the index of the 'DOT' column
    for row in sheet.iter_rows(min_row=startrow + start_distance + 1, max_row=sheet.max_row, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.font = font
            cell.border = border
            cell.alignment = alignment
            if frequency:
                cell.number_format = '0.00%'
            elif cell.column == dot_index:
                cell.number_format = '0.0'
            else:
                cell.number_format = '0.00'  # 2 decimal places for other columns

    book.save(filename)
    return startrow + start_distance + 1, len(df.columns)


def append_chart_to_excel_openpy(y_axis_title, name, filename, startrow, len_df, column, sheet_name, cropped,flip_x_y=False):
    # Step 1: Load the existing workbook and worksheet
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Step 5: Create a BarChart object
    chart = BarChart()
    chart.title = name
    chart.legend.position = 'b'
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.line.noFill = True
    chart.graphical_properties.line.prstDash = None
    chart.width = 21
    h = 0.8
    w = 0.8
    x = 0
    y = 0
    hl = None
    if flip_x_y:
        h = 0.4
        w = 0.8
        x = 0
        y = 0
        hl = 0.25

    chart.layout = Layout(
        ManualLayout(
            x=x, y=y,
            h=h, w=w
        )
    )
    chart.legend.layout = Layout(
        ManualLayout(
            x=0, y=1,
            w=1,h=hl
        )
    )

    chart_title_font = openpyxl.drawing.text.Font(typeface='Calibri')
    chart_title_props = CharacterProperties(
        sz=1400,  # Size in EMU (14pt * 100)
        b=True,  # Bold
        u="sng",  # Underline (single)
        solidFill="FF0000"  # Red color in hex (FF0000)
    )

    # Apply the font settings to the title
    chart.title.tx.rich.p[0].r[0].rPr = chart_title_props
    chart.title.tx.rich.p[0].r[0].rPr.latin = chart_title_font
    # Step 6: Define data for the chart
    if not cropped:
        startrow = startrow + 1
    from_rows = False
    if flip_x_y:
        from_rows = True
        data_range = Reference(ws, min_col=1, min_row=startrow, max_col=len_df, max_row=ws.max_row)
        categories = Reference(ws, min_col=2, min_row=startrow -1, max_row=startrow -1, max_col=len_df)
    else:
        data_range = Reference(ws, min_col=column + 1, min_row=startrow -1, max_col=len_df, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=startrow, max_row=ws.max_row)

    # Add data and categories to the chart
    chart.add_data(data_range, from_rows=from_rows,titles_from_data=True)
    chart.set_categories(categories)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.title = ""
    if cropped:
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        y_axis_title = "%"
    # y-title text properties
    yt_color = "ff0000"

    xml = f"""
    <txPr>
      <a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        <a:r>
           <a:rPr b="1" i="0" sz="1000" spc="-1" strike="noStrike">
              <a:solidFill>
                 <a:srgbClr val="{yt_color}" />
              </a:solidFill>
              <a:latin typeface="Calibri" />
           </a:rPr>
           <a:t>{y_axis_title}</a:t>
        </a:r>
      </a:p>
    </txPr>
    """

    chart.y_axis.title.tx.rich = RichText.from_tree(openpyxl.xml.functions.fromstring(xml))
    # Step 7: Insert the chart into the worksheet
    AsciiOfLetter = ord('A')
    number_of_lines_to_add = 2
    if cropped:
        number_of_lines_to_add += 1
    IncrementedLettersAscii = AsciiOfLetter + len_df + number_of_lines_to_add
    ws.add_chart(chart, f"{chr(IncrementedLettersAscii)}{startrow -1}")
    # Step 8: Save the Excel file
    wb.save(filename)


def write_text_to_excel(filename, text, sheet_name='Sheet1'):
    # Try to open an existing workbook
    try:
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    font = Font(name='David', size=12, bold=True)
    # Find the next empty row
    start_row = ws.max_row + 2
    # Write the text to the first cell in the next empty row
    cell = ws.cell(row=start_row, column=1, value=text)
    cell.font = font
    wb.save(filename)


# Function to find the first row with the date format
def find_first_date_row(df):
    for index, row in df.iterrows():
        for col in row:
            try:
                pd.to_datetime(col, format='%m/%d/%y')
                return index
            except (ValueError, TypeError):
                continue
    else:
        return -1
