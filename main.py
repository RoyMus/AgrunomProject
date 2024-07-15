from pathlib import Path
import pandas as pd
from customtkinter import *
from tkinter import filedialog, ttk
from CalcUtils import *
import statsmodels.api as sm
from statsmodels.formula.api import ols
from scipy import stats
import re
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import shutil


# Define a custom sorting key function
def custom_sort_key(val):
    if val == 'UTC':
        return (0, '')  # Make "UTC" come first
    return (1, int(val))  # Then sort by numeric values


class ScrollableCheckBoxFrame(CTkScrollableFrame):
    def __init__(self, master, item_list, command=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        self.checkbox_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        checkbox = CTkCheckBox(self, text=item)
        if self.command is not None:
            checkbox.configure(command=self.command)
        checkbox.grid(row=len(self.checkbox_list), column=0, pady=(0, 10))
        self.checkbox_list.append(checkbox)

    def remove_item(self, item):
        for checkbox in self.checkbox_list:
            if item == checkbox.cget("text"):
                checkbox.destroy()
                self.checkbox_list.remove(checkbox)
                return

    def get_checked_items(self):
        return [checkbox.cget("text") for checkbox in self.checkbox_list if checkbox.get() == 1]


class ScrollableRadiobuttonFrame(CTkScrollableFrame):
    def __init__(self, master, item_list, command=None, text_variable=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        if text_variable is not None:
            self.radiobutton_variable = text_variable
        else:
            self.radiobutton_variable = StringVar()

        self.radiobutton_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        radiobutton = CTkRadioButton(self, text=item, value=item, variable=self.radiobutton_variable)
        if self.command is not None:
            radiobutton.configure(command=self.command)
        radiobutton.grid(row=len(self.radiobutton_list), column=0, pady=(0, 10))
        self.radiobutton_list.append(radiobutton)

    def remove_item(self, item):
        for radiobutton in self.radiobutton_list:
            if item == radiobutton.cget("text"):
                radiobutton.destroy()
                self.radiobutton_list.remove(radiobutton)
                return

    def get_checked_item(self):
        return self.radiobutton_variable.get()


def reverse_hebrew_sentence(sentence):
    # Regular expression to match Hebrew words, possibly followed by numbers or special characters
    pattern = re.compile(r'([א-ת]+(?:\d+|[^\sא-ת]*)*)')

    # Find all matches
    words_numbers = pattern.findall(sentence)

    # Reverse the list of matches
    reversed_sentence = ' '.join(words_numbers[::-1])

    return reversed_sentence


def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    path = Path(filename)
    output_file = path.with_stem(f"{path.stem}(app_generated)")
    if not Path(output_file).is_file():
        output_path = shutil.copy(filename, output_file)
    else:
        output_path = output_file
    # Try to open an existing workbook
    try:
        book = load_workbook(output_path)
        writer = pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.workbook = book
        if sheet_name in writer.workbook.sheetnames:
            startrow = writer.workbook[sheet_name].max_row
        else:
            startrow = 0
    except FileNotFoundError:
        # File does not exist yet, create a new one
        writer = pd.ExcelWriter(output_path, engine='openpyxl', mode='w')
        startrow = 0
    # Write the dataframe to the Excel file
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow + 1, index=False, header=True)
    writer.close()
    # Reload the workbook and the sheet to apply formatting
    book = load_workbook(output_path)
    sheet = book[sheet_name]

    # Apply font and borders
    font = Font(name='David', size=12)
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))
    alignment = Alignment(horizontal='center')

    for row in sheet.iter_rows(min_row=startrow + 2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = font
            cell.border = border
            cell.alignment = alignment
    book.save(output_path)


class AgrunomProjectApplication(CTkFrame):
    def __init__(self, root):
        self.filename = None
        set_appearance_mode("light")  # Modes: system (default), light, dark
        set_default_color_theme("green")
        # Themes: blue (default), dark-blue, green
        CTkFrame.__init__(self, root)
        root.columnconfigure(0, weight=1)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(0, weight=1)
        root.rowconfigure(1, weight=1)
        self.upper_frame = CTkFrame(root, fg_color="transparent")
        self.upper_frame.grid(row=0, column=0, columnspan=2)
        self.scrollable_checkbox_frame = None
        self.value_inside = None
        self.input_df = None
        self.output_df = None
        self.root = root
        read_file_button = CTkButton(self.upper_frame, text="Open File", command=self.read_file)
        read_file_button.grid(row=0, column=0, columnspan=2, pady=10)
        # create tabview
        self.tabview = CTkTabview(self.upper_frame, width=250)
        self.tabview.add("X")
        self.tabview.add("Y")
        self.tabview.add("Block")
        self.tabview.tab("X").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        self.tabview.tab("Y").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Block").grid_columnconfigure(0, weight=1)
        self.excel_frame = CTkScrollableFrame(self.root, label_text="Excel Data", orientation="horizontal")
        self.tv1 = ttk.Treeview(self.excel_frame)
        self.tv1.grid(row=0, column=0, sticky="nsew")
        self.excel_frame.columnconfigure(0, weight=1)
        self.excel_frame.rowconfigure(0, weight=1)

    def calculate_students_t(self):
        x_value = self.x.get_checked_item()
        block = self.block_selection.get_checked_item()
        items = self.scrollable_checkbox_frame.get_checked_items()
        self.output_dict = {}
        output_columns = ["DOT", "DOT sig", "3DAT", "3DAT sig", "7DAT", "7DAT sig", "10DAT", "10DAT sig", "14DAT",
                          "14DAT sig"]
        index = 0
        for label in items:
            for col in self.input_df.columns:
                if label != col.split('.')[0]:
                    continue
                df = pd.DataFrame({'Treatment': self.input_df[x_value], 'Value': self.input_df[col]})
                df['Value'] = pd.to_numeric(df['Value'])
                df['Treatment'] = df['Treatment'].astype(str)
                if block != "None":
                    df["Block"] = self.input_df[block]
                    model = ols('Value ~ C(Treatment) + C(Block)', data=df).fit()
                else:
                    model = ols('Value ~ C(Treatment)', data=df).fit()

                # Perform ANOVA
                anova_table = sm.stats.anova_lm(model, typ=2)

                # Calculate Mean Squared Error (MSE)
                mse = anova_table['sum_sq']['Residual'] / anova_table['df']['Residual']

                # Number of samples per group (assuming equal sample size)
                n = df['Treatment'].value_counts().min()  # Use the minimum count in case of unequal sizes

                # Degrees of freedom for error
                df_error = anova_table['df']['Residual']

                # Significance level
                alpha = 0.05

                # Critical t-value
                t_critical = stats.t.ppf(1 - alpha / 2, df_error)

                # Calculate means for each treatment
                treatment_means = df.groupby('Treatment')['Value'].mean().to_dict()
                treatment_means = {r: treatment_means[r] for r in
                                   sorted(treatment_means, key=treatment_means.get, reverse=True)}
                sigByKey = calculate_significant_letters(treatment_means, t_critical, mse, n)
                index = index % len(output_columns)
                self.output_dict[output_columns[index]] = list(treatment_means.values())
                self.output_dict[output_columns[index + 1]] = ["".join(sorted(value)) for value in sigByKey.values()]
                index += 2

            self.output_df = pd.DataFrame(self.output_dict)
            self.output_df.insert(0, label, treatment_means.keys())
            self.output_df = self.output_df.sort_values(by=label, key=lambda col: col.map(custom_sort_key))
            append_df_to_excel(self.filename, self.output_df, "טבלאות")

    def calculate_tukey_hsd(self):
        x_value = self.x.get_checked_item()
        block = self.block_selection.get_checked_item()
        for label in self.scrollable_checkbox_frame.get_checked_items():
            df = pd.DataFrame({'Treatment': self.input_df[x_value], 'Value': self.input_df[label]})
            print("DataFrame:")
            print(df)
            df['Value'] = pd.to_numeric(df['Value'])
            df['Treatment'] = df['Treatment'].astype(str)
            df = df.dropna()
            if block != "None":
                df["Block"] = self.df[block]
                model = ols('Value ~ C(Treatment) + C(Block)', data=df).fit()
            else:
                model = ols('Value ~ C(Treatment)', data=df).fit()

            # Perform ANOVA
            anova_table = sm.stats.anova_lm(model, typ=2)

            # Perform Tukey HSD test
            tukey = pairwise_tukeyhsd(endog=df['Value'],
                                      groups=df['Treatment'],
                                      alpha=0.05)
            print(tukey)

    def read_file(self):
        self.filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                   filetypes=[("Excel files", ".xlsx .xls")])
        if not self.filename:
            return
        with open(self.filename, 'rb') as f:
            self.excel_frame._label.configure(text=reverse_hebrew_sentence(Path(self.filename).stem))
            # self.clear_data()
            self.input_df = pd.read_excel(f, "תוצאות", index_col=False)
            columns = list(self.input_df.columns)
            self.tv1["column"] = columns
            self.tv1["show"] = "headings"
            for column in self.tv1["column"]:
                self.tv1.heading(column, text=column)
            df_rows = self.input_df.to_numpy().tolist()
            for row in df_rows:
                self.tv1.insert("", "end", values=row)
            self.input_df = self.input_df.drop(self.input_df.index[-1])
            self.tabview.grid(row=1, column=0, columnspan=2)
            variable = StringVar()
            variable.set("טיפול")
            unduplicatedcolumns = []
            for col in columns:
                if col.split('.')[0] not in unduplicatedcolumns:
                    unduplicatedcolumns.append(col)
            self.x = ScrollableRadiobuttonFrame(master=self.tabview.tab("X"), item_list=unduplicatedcolumns,
                                                text_variable=variable)
            self.x.grid(row=0, column=0)
            self.scrollable_checkbox_frame = ScrollableCheckBoxFrame(master=self.tabview.tab("Y"),
                                                                     item_list=unduplicatedcolumns)
            self.scrollable_checkbox_frame.grid(row=0, column=0)
            calc_button = CTkButton(self.upper_frame, text="Calculate Each Pair Student's T's",
                                    command=self.calculate_students_t)
            self.step_variable = StringVar()
            self.step_variable.set("None")
            blocks = unduplicatedcolumns
            blocks.insert(0, "None")
            self.block_selection = ScrollableRadiobuttonFrame(master=self.tabview.tab("Block"), item_list=blocks,
                                                              text_variable=self.step_variable)
            self.block_selection.grid(row=0, column=0)
            calc_button.grid(row=2, column=0, pady=(15, 0), padx=10)
            self.excel_frame.grid(row=1, column=0, sticky="nsew", columnspan=2)
            tukey_calc = CTkButton(self.upper_frame, text="Calculate Each Pair Tukey ",
                                   command=self.calculate_tukey_hsd)
            tukey_calc.grid(row=2, column=1, pady=(15, 0), padx=10)
            # calc_button = CTkButton(self.root, text="Calculate Each Pair Tukey", command=self.print_table)
            # calc_button.grid(row=2 + n_cols + 1, column=1,padx=5)

    def clear_data(self):
        self.tv1.delete(*self.tv1)


if __name__ == '__main__':
    root = CTk()
    root.geometry("500x700")
    root.title('AgronumProject')
    AgrunomProjectApplication(root)
    root.mainloop()
